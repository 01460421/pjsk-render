#!/usr/bin/env python3
"""
PJSK 私車管理機器人 — 主系統
所有輸出採用圖像渲染（PingFang TC 蘋方風格）
拆分架構: car_bot.py (主程式) + img_render.py (圖像引擎)
"""
import discord
from discord import app_commands
from discord.ui import Button, View, Select, Modal, TextInput
import pandas as pd
from typing import Dict, List, Optional
import os, asyncio, json, re, random, math, csv, zipfile, shutil
from io import BytesIO, StringIO
from datetime import datetime, timedelta, date
from aiohttp import ClientSession, ClientTimeout
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 載入 .env (雲端部署用)
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

from img_render import (
    render_table_image, render_info_card, render_message_box,
    render_help_image, render_line_chart, Theme, CJK_FONT, SERIF_FONT
)
from render_funcs import (
    calc_song_score, calc_ep_value, find_push_plans,
    create_push_plan_image, create_ranking_detail_image,
    create_ranking_list_image, create_ranking_chart as _local_ranking_chart,
    create_schedule_image as _local_schedule_image,
    create_member_table_image, create_hours_table_image,
    ENERGY_MULTIPLIERS, SONG_DB, load_song_db
)

# ========== 常數 ==========
RENDER_URL = os.getenv('RENDER_URL', '')  # 雲端渲染 URL (空=本地渲染)
RENDER_API_KEY = os.getenv('RENDER_API_KEY', '')
HISEKAI_API = "https://api.hisekai.org"
PJSK_CENTER = "https://project-sekai-center.vercel.app"
CAR_TYPES = ["蝦","臉","sage","10th","任意","高難","雪初音"]
TIME_SLOTS = [f"{h:02d}:00" for h in range(24)]
TRACKED_RANKS = [1,2,3,10,20,50,100]
RANKING_HISTORY_FILE = "ranking_history.json"
DATA_FILE = "pjsk_car_data.json"
ADMIN_ROLE_ID = 1438186385386377267  # 管理員身份組 ID

# ========== 持久化 ==========
def load_json(path, default):
    if os.path.exists(path):
        try:
            with open(path,'r',encoding='utf-8') as f: return json.load(f)
        except: pass
    return default

def save_json(path, data):
    with open(path,'w',encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

bot_data = load_json(DATA_FILE, {
    "members":{},"schedule":{},"rewards":{},"rooms":{},
    "settings":{"registration_open":True,"schedule_open":False},"stats":{}
})
ranking_history = load_json(RANKING_HISTORY_FILE, {"event_name":"","records":[]})

def save_data():   save_json(DATA_FILE, bot_data)
def save_ranking(): save_json(RANKING_HISTORY_FILE, ranking_history)

# ========== 遠端渲染代理 ==========
async def _remote_render(func_name, **kwargs):
    """嘗試遠端渲染，失敗回傳 None (降級本地)"""
    if not RENDER_URL:
        return None
    try:
        headers = {'Content-Type': 'application/json'}
        if RENDER_API_KEY:
            headers['X-API-Key'] = RENDER_API_KEY
        # 序列化 (處理不可序列化的物件)
        clean = json.loads(json.dumps(kwargs, default=str))
        payload = json.dumps({'func': func_name, 'kwargs': clean})
        async with ClientSession() as s:
            async with s.post(f"{RENDER_URL}/render",
                              data=payload, headers=headers,
                              timeout=ClientTimeout(total=30)) as resp:
                if resp.status == 200:
                    data = await resp.read()
                    return BytesIO(data)
                else:
                    err = await resp.text()
                    print(f"[remote_render] {func_name} failed ({resp.status}): {err[:200]}")
    except Exception as e:
        print(f"[remote_render] {func_name} error: {e}")
    return None

# 班表圖 (包裝: 注入 members 資料)
def create_schedule_image(dt, schedule, dpi=130):
    members = bot_data.get("members", {})
    return _local_schedule_image(dt, schedule, members=members, dpi=dpi, pjsk_center=PJSK_CENTER)

# 排名走勢圖 (包裝: 注入 ranking_history)
def create_ranking_chart(rank=None, event_name=None):
    records = ranking_history.get("records", [])
    if not event_name:
        ce = ranking_history.get("event_name", "")
        if ce:
            records = [r for r in records if r.get("event", r.get("time","")) == ce or "event" not in r]
    return _local_ranking_chart(records, rank=rank, event_name=event_name)

# ========== 工具 ==========
def is_admin(interaction: discord.Interaction) -> bool:
    """檢查是否有管理員身份組"""
    if not interaction.guild: return False
    return any(r.id == ADMIN_ROLE_ID for r in interaction.user.roles)

def admin_check():
    """裝飾器: 限制指令僅管理員可用"""
    async def predicate(interaction: discord.Interaction) -> bool:
        if is_admin(interaction):
            return True
        await interaction.response.send_message("此指令僅限管理員使用", ephemeral=True,silent=True)
        return False
    return app_commands.check(predicate)

def fmt_num(n):
    if not n: return "-"
    return f"{n/10000:.2f}萬" if n >= 10000 else f"{int(n):,}"

def get_today(): return datetime.now().strftime("%Y-%m-%d")

def parse_time_range(s):
    m = re.match(r'(\d{1,2})-(\d{1,2})', s)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        if b <= a: b += 24
        return [f"{h%24:02d}:00" for h in range(a,b)]
    return []

def calculate_bonus(leader, members):
    return round((leader + 100 + sum(members)/5) / 100, 2)

# ========== ScoreTable ==========
class ScoreTable:
    def __init__(self, xlsx_path):
        df = pd.read_excel(xlsx_path, header=None)
        self.score_map = {}
        for r in range(3, len(df)):
            mn, mx = df.iloc[r,2], df.iloc[r,4]
            if pd.isna(mn) or pd.isna(mx) or int(mx)>1019999: continue
            rng = f"{int(mn)}~{int(mx)}"
            for ci, bonus in enumerate([float(x) for x in df.iloc[2,5:].dropna().tolist()]):
                if bonus > 2.50: continue
                base = df.iloc[r, 5+ci]
                if pd.isna(base): continue
                for e, mult in ENERGY_MULTIPLIERS.items():
                    actual = int(base)*mult
                    if actual not in self.score_map:
                        self.score_map[actual] = (rng, bonus, e)
        self.scores = sorted(self.score_map.keys(), reverse=True)
        self.score_set = set(self.scores)
        print(f"[ScoreTable] {len(self.scores)} values loaded")

def find_solution(tbl, target, max_plays=50):
    if not tbl or target<=0: return None
    def ms(s,p):
        o=tbl.score_map[s]
        return {'range':o[0],'bonus':o[1],'energy':o[2],'score':s,'plays':p,'total':s*p}
    for s in tbl.scores:
        if s<=target and target%s==0 and target//s<=max_plays: return [ms(s,target//s)]
    for s1 in tbl.scores:
        if s1>target: continue
        for p1 in range(min(target//s1, max_plays), 0, -1):
            rem=target-s1*p1
            if rem==0: return [ms(s1,p1)]
            if rem in tbl.score_set: return [ms(s1,p1), ms(rem,1)]
    return None

def create_schedule_excel(dt, schedule):
    """生成班表 Excel 檔"""
    wb = Workbook()
    ws = wb.active
    ws.title = "私車班表"
    
    # 樣式定義
    title_font = Font(name='PingFang TC', size=16, bold=True, color='1A1A2E')
    header_font = Font(name='PingFang TC', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1B2838', end_color='1B2838', fill_type='solid')
    p2_header_fill = PatternFill(start_color='8E44AD', end_color='8E44AD', fill_type='solid')
    data_font = Font(name='PingFang TC', size=10, color='1A1A2E')
    bonus_font = Font(name='PingFang TC', size=10, bold=True, color='C0392B')
    p1_font = Font(name='PingFang TC', size=10, bold=True, color='E67E22')
    p2_font = Font(name='PingFang TC', size=10, bold=True, color='8E44AD')
    even_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    active_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    footer_font = Font(name='PingFang TC', size=9, color='6C7A89', italic=True)
    thin_border = Border(
        left=Side(style='thin', color='D5D8DC'), right=Side(style='thin', color='D5D8DC'),
        top=Side(style='thin', color='D5D8DC'), bottom=Side(style='thin', color='D5D8DC'))
    center = Alignment(horizontal='center', vertical='center')
    
    # 標題
    ws.merge_cells('A1:J1')
    c = ws['A1']
    c.value = f"私車班表 — {dt}"
    c.font = title_font
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    
    # 表頭
    headers = ["時間","車種","平均倍率","P1","P2(S6)","P3","P4","P5","外援","備註"]
    col_widths = [8, 8, 10, 8, 28, 24, 24, 24, 24, 14]
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = header_font
        cell.fill = p2_header_fill if i == 5 else header_fill
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[chr(64+i)].width = w
    ws.row_dimensions[2].height = 24
    
    # 資料列
    members = bot_data.get("members", {})
    def get_bonus(p):
        if not p: return 0
        b = p.get('bonus', 0) or 0
        if b == 0 and p.get('user_id') and p['user_id'] in members:
            b = members[p['user_id']].get('bonus', 0)
        return b
    def fp(p):
        if not p: return ""
        b = get_bonus(p)
        name = p.get('name','')
        return f"{name}({b:.2f})" if b > 0 else name
    def fs6(p):
        if not p: return ""
        n=p.get("name",""); b=get_bonus(p)
        pw=p.get("s6_power") or p.get("power",0) or 0
        if pw==0 and p.get('user_id') and p['user_id'] in members:
            pw=members[p['user_id']].get('s6_power',0) or members[p['user_id']].get('power',0)
        if b > 0:
            return f"{n}({b:.2f}/{pw/10000:.2f}萬)" if pw>0 else f"{n}({b:.2f})"
        else:
            return f"{n}({pw/10000:.2f}萬)" if pw>0 else n
    
    for ri, hour in enumerate(TIME_SLOTS):
        row = ri + 3
        sh = schedule.get(hour, {})
        has_data = sh.get("p2") or sh.get("p3") or sh.get("p4") or sh.get("p5")
        
        vals = [hour, sh.get("car_type","蝦"),
                f"{sh.get('avg_bonus',0):.2f}" if sh.get('avg_bonus') else "",
                "omega", fs6(sh.get("p2")), fp(sh.get("p3")),
                fp(sh.get("p4")), fp(sh.get("p5")),
                fp(sh.get("support")), sh.get("note","")]
        
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.border = thin_border
            cell.alignment = center
            # 字體
            if ci == 3 and val:  # 倍率
                cell.font = bonus_font
            elif ci == 4:  # P1
                cell.font = p1_font
            elif ci == 5:  # P2
                cell.font = p2_font
            else:
                cell.font = data_font
            # 背景
            if has_data:
                cell.fill = active_fill
            elif ri % 2 == 0:
                cell.fill = even_fill
        
        ws.row_dimensions[row].height = 22
    
    # 底部
    footer_row = len(TIME_SLOTS) + 4
    ws.merge_cells(f'A{footer_row}:J{footer_row}')
    c = ws[f'A{footer_row}']
    c.value = f"P1: omega | P2: S6 | P3–P5: 推手 | {PJSK_CENTER} | {datetime.now().strftime('%H:%M:%S')}"
    c.font = footer_font
    c.alignment = Alignment(horizontal='center')
    
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ========== 排班邏輯 ==========
def is_signup_closed(hour_str):
    now=datetime.now()
    try: sh=int(hour_str.split(":")[0])
    except: return (False,"")
    st=now.replace(hour=sh,minute=0,second=0,microsecond=0)
    if st<now-timedelta(hours=12): st+=timedelta(days=1)
    hu=(st-now).total_seconds()/3600
    today=get_today()
    sd=bot_data.get("schedule",{}).get(today,{}).get(hour_str,{})
    has_s6=sd.get("p2") is not None
    if hu<=1: return (True,f"{hour_str} 已截止（前1h）")
    if hu<=2 and has_s6: return (True,f"{hour_str} 已截止（前2h，S6到位）")
    return (False,"")

def auto_assign_schedule(dt, hour, applicants):
    if not applicants: return {}
    def sk(x):
        b=x.get('bonus',0); t=x.get('registered_at','')
        return (-round(b*50)/50, t)
    sa=sorted(applicants, key=sk)
    shift={"car_type":"蝦","p1":{"name":"omega","fixed":True},
           "p2":None,"p3":None,"p4":None,"p5":None,"support":None,"avg_bonus":0,"note":""}
    s6a=[a for a in sa if a.get('role')=='s6']
    spa=[a for a in sa if a.get('role')=='support']
    psa=[a for a in sa if a.get('role') not in ['s6','support']]
    if s6a:
        s6=s6a[0].copy()
        if s6.get('s6_bonus',0)>0: s6['bonus']=s6['s6_bonus']
        shift["p2"]=s6
    if spa: shift["support"]=spa[0]
    positions=["p3","p4","p5"]; assigned=0
    for app in psa:
        if assigned>=3: break
        multi=app.get('multi','單開'); accs={'單開':1,'雙開':2,'三開':3}.get(multi,1)
        ab=[app.get('bonus',0),app.get('bonus_2',0) or app.get('bonus',0),app.get('bonus_3',0) or app.get('bonus',0)]
        for ai in range(min(accs,3-assigned)):
            c=app.copy(); c['bonus']=ab[ai]
            if ai>0: c['name']=f"{app.get('name','')}({ai+1}開)"
            shift[positions[assigned]]=c; assigned+=1
    bs=[shift[p].get('bonus',0) for p in ["p3","p4","p5"] if shift[p]]
    shift["avg_bonus"]=sum(bs)/len(bs) if bs else 0
    return shift

def refresh_schedule(dt=None):
    if dt is None: dt=get_today()
    schedule=bot_data.get("schedule",{}).get(dt,{})
    for h in list(schedule.keys()):
        apps=schedule[h].get("applicants",[])
        if apps:
            shift=auto_assign_schedule(dt,h,apps)
            shift["applicants"]=apps; schedule[h].update(shift)
    save_data()

# ========== Discord Bot ==========
table: Optional[ScoreTable] = None
intents = discord.Intents.default()
intents.message_content = True; intents.members = True
client = discord.Client(intents=intents)
tree = app_commands.CommandTree(client)

grp_member   = app_commands.Group(name="成員", description="成員管理")
grp_schedule = app_commands.Group(name="班表", description="報班系統")
grp_room     = app_commands.Group(name="房間", description="房間管理")
grp_reward   = app_commands.Group(name="獎勵", description="獎勵系統")
grp_query    = app_commands.Group(name="查詢", description="查詢功能")
grp_tools    = app_commands.Group(name="工具", description="實用工具")
grp_system   = app_commands.Group(name="系統", description="[管理員] 系統管理")

# ========== 互動 UI ==========
class SignupScheduleModal(Modal, title="報班"):
    time_input = TextInput(label="時段 (如 08-12)",placeholder="08-12",required=True,max_length=10)
    note_input = TextInput(label="備註",placeholder="可留空",required=False,max_length=50)
    def __init__(self,role="pusher",multi="單開"):
        super().__init__(); self._role=role; self._multi=multi
    async def on_submit(self, interaction):
        uid=str(interaction.user.id)
        if uid not in bot_data.get("members",{}):
            img=render_message_box("錯誤",["請先 /成員 註冊"],accent_color=Theme.RED)
            await interaction.response.send_message(file=discord.File(img,"err.png"),ephemeral=True,silent=True); return
        if not bot_data.get("settings",{}).get("schedule_open"):
            img=render_message_box("錯誤",["報班未開放"],accent_color=Theme.RED)
            await interaction.response.send_message(file=discord.File(img,"err.png"),ephemeral=True,silent=True); return
        hours=parse_time_range(self.time_input.value.strip())
        if not hours:
            img=render_message_box("錯誤",["格式: 08-12"],accent_color=Theme.RED)
            await interaction.response.send_message(file=discord.File(img,"err.png"),ephemeral=True,silent=True); return
        today=get_today(); bot_data.setdefault("schedule",{}).setdefault(today,{})
        closed_hours=[]; open_hours=[]
        for h in hours:
            closed,reason=is_signup_closed(h)
            if closed: closed_hours.append(reason)
            else: open_hours.append(h)
        if not open_hours:
            img=render_message_box("已截止",closed_hours,accent_color=Theme.RED)
            await interaction.response.send_message(file=discord.File(img,"x.png"),ephemeral=True,silent=True); return
        m=bot_data["members"][uid]
        app={"user_id":uid,"name":m["name"],"bonus":m["bonus"],"bonus_2":m.get("bonus_2",0),
             "bonus_3":m.get("bonus_3",0),"s6_bonus":m.get("s6_bonus",0),"power":m["power"],
             "s6_power":m.get("s6_power",0),"multi":self._multi,"role":self._role,
             "note":self.note_input.value.strip(),"registered_at":datetime.now().isoformat()}
        registered=[]
        for h in open_hours:
            if h not in bot_data["schedule"][today]: bot_data["schedule"][today][h]={"applicants":[]}
            if not any(a["user_id"]==uid for a in bot_data["schedule"][today][h].get("applicants",[])):
                bot_data["schedule"][today][h].setdefault("applicants",[]).append(app); registered.append(h)
        save_data()
        if registered:
            refresh_schedule(today)
            rn={"pusher":"推手","s6":"S6","support":"外援"}.get(self._role,self._role)
            lines=[f"時段: {', '.join(registered)}",f"角色: {rn}",f"開數: {self._multi}",f"倍率: {m['bonus']:.2f}"]
            img=render_message_box("報班成功",lines,accent_color=Theme.GREEN)
            await interaction.response.send_message(file=discord.File(img,"ok.png"),silent=True)
        else:
            img=render_message_box("提示",["這些時段已報過"],accent_color=Theme.ORANGE)
            await interaction.response.send_message(file=discord.File(img,"dup.png"),ephemeral=True,silent=True)

class EditScheduleModal(Modal, title="編輯班表"):
    hour_input = TextInput(label="時段 (如 08:00 或 08-12)",placeholder="08-12",required=True,max_length=10)
    s6_input = TextInput(label="P2/S6 名稱",placeholder="S6玩家",required=False)
    p3_input = TextInput(label="P3 名稱",placeholder="推手",required=False)
    p4_input = TextInput(label="P4 名稱",placeholder="推手",required=False)
    p5_input = TextInput(label="P5 名稱",placeholder="推手",required=False)
    async def on_submit(self, interaction):
        raw=self.hour_input.value.strip(); today=get_today()
        bot_data.setdefault("schedule",{}).setdefault(today,{})
        # 支援範圍 (08-12) 或單一時段 (08:00)
        hours=parse_time_range(raw)
        if not hours:
            # 嘗試單一時段
            if re.match(r'\d{1,2}:\d{2}', raw):
                hours = [raw]
            elif re.match(r'\d{1,2}', raw):
                hours = [f"{int(raw):02d}:00"]
            else:
                img=render_message_box("錯誤",["格式: 08-12 或 08:00"],accent_color=Theme.RED)
                await interaction.response.send_message(file=discord.File(img,"e.png"),ephemeral=True,silent=True); return
        def fm(name):
            if not name: return None
            for uid,m in bot_data["members"].items():
                if m.get("name","").lower()==name.lower() or name.lower() in m.get("name","").lower():
                    return {"user_id":uid,**m}
            return {"name":name,"bonus":0,"power":0,"s6_power":0}
        p2=fm(self.s6_input.value.strip()); p3=fm(self.p3_input.value.strip())
        p4=fm(self.p4_input.value.strip()); p5=fm(self.p5_input.value.strip())
        changed = []
        for hour in hours:
            if hour not in bot_data["schedule"][today]:
                bot_data["schedule"][today][hour]={"car_type":"蝦","p1":{"name":"omega","fixed":True},
                    "p2":None,"p3":None,"p4":None,"p5":None,"support":None,"avg_bonus":0,"note":"","applicants":[]}
            sh=bot_data["schedule"][today][hour]
            # 只覆蓋有填的欄位，留空保留原本
            if p2 is not None: sh["p2"]=p2
            if p3 is not None: sh["p3"]=p3
            if p4 is not None: sh["p4"]=p4
            if p5 is not None: sh["p5"]=p5
            bs=[sh[k].get('bonus',0) for k in ["p3","p4","p5"] if sh.get(k)]
            sh["avg_bonus"]=sum(bs)/len(bs) if bs else 0
        save_data()
        range_str = f"{hours[0]}~{hours[-1]}" if len(hours)>1 else hours[0]
        # 顯示最終狀態 (取最後一個時段)
        last=bot_data["schedule"][today][hours[-1]]
        img=render_info_card("班表已更新",[("時段",f"{range_str} ({len(hours)}h)"),
            ("P2/S6",last['p2']['name'] if last.get('p2') else '-'),
            ("P3",last['p3']['name'] if last.get('p3') else '-'),
            ("P4",last['p4']['name'] if last.get('p4') else '-'),
            ("P5",last['p5']['name'] if last.get('p5') else '-'),
            ("備註","留空欄位已保留原設定")],accent_color=Theme.GREEN)
        await interaction.response.send_message(file=discord.File(img,"edit.png"),silent=True)

class ScheduleView(View):
    def __init__(self): super().__init__(timeout=300)
    @discord.ui.button(label="推手報班",style=discord.ButtonStyle.success,emoji="🎯",row=0)
    async def pusher_btn(self, interaction, button):
        if not bot_data.get("settings",{}).get("schedule_open"):
            await interaction.response.send_message("報班未開放",ephemeral=True,silent=True); return
        await interaction.response.send_modal(SignupScheduleModal("pusher","單開"))
    @discord.ui.button(label="S6報班",style=discord.ButtonStyle.primary,emoji="⭐",row=0)
    async def s6_btn(self, interaction, button):
        if not bot_data.get("settings",{}).get("schedule_open"):
            await interaction.response.send_message("報班未開放",ephemeral=True,silent=True); return
        await interaction.response.send_modal(SignupScheduleModal("s6","單開"))
    @discord.ui.button(label="雙開",style=discord.ButtonStyle.secondary,emoji="2️⃣",row=0)
    async def dual_btn(self, interaction, button):
        await interaction.response.send_modal(SignupScheduleModal("pusher","雙開"))
    @discord.ui.button(label="三開",style=discord.ButtonStyle.secondary,emoji="3️⃣",row=0)
    async def tri_btn(self, interaction, button):
        await interaction.response.send_modal(SignupScheduleModal("pusher","三開"))
    @discord.ui.button(label="編輯",style=discord.ButtonStyle.primary,emoji="✏️",row=1)
    async def edit_btn(self, interaction, button):
        await interaction.response.send_modal(EditScheduleModal())
    @discord.ui.button(label="重新整理",style=discord.ButtonStyle.secondary,emoji="🔄",row=1)
    async def refresh_btn(self, interaction, button):
        await interaction.response.defer()
        today=get_today(); schedule=bot_data.get("schedule",{}).get(today,{})
        xlsx=create_schedule_excel(today, schedule)
        await interaction.followup.send(file=discord.File(xlsx,f"班表_{today}.xlsx"),view=ScheduleView(),silent=True)
    @discord.ui.button(label="放大圖片",style=discord.ButtonStyle.secondary,emoji="🔍",row=1)
    async def zoom_btn(self, interaction, button):
        await interaction.response.defer()
        today=get_today(); schedule=bot_data.get("schedule",{}).get(today,{})
        img=create_schedule_image(today, schedule, dpi=200)
        await interaction.followup.send(file=discord.File(img,"schedule_hd.png"),ephemeral=True,silent=True)
    @discord.ui.button(label="Excel",style=discord.ButtonStyle.secondary,emoji="📊",row=1)
    async def excel_btn(self, interaction, button):
        await interaction.response.defer()
        today=get_today(); schedule=bot_data.get("schedule",{}).get(today,{})
        xlsx=create_schedule_excel(today, schedule)
        await interaction.followup.send(file=discord.File(xlsx,f"班表_{today}.xlsx"),ephemeral=True,silent=True)

# ========== /help ==========
def render_help_excel(sections, link=""):
    """生成指令手冊 Excel 檔"""
    wb = Workbook()
    ws = wb.active
    ws.title = "指令手冊"
    
    # 樣式
    title_font = Font(name='PingFang TC', size=16, bold=True, color='1A1A2E')
    section_font = Font(name='PingFang TC', size=12, bold=True, color='8E44AD')
    section_fill = PatternFill(start_color='F3E5F5', end_color='F3E5F5', fill_type='solid')
    cmd_font = Font(name='PingFang TC', size=11, bold=True, color='2980B9')
    desc_font = Font(name='PingFang TC', size=11, color='1A1A2E')
    header_font = Font(name='PingFang TC', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1B2838', end_color='1B2838', fill_type='solid')
    link_font = Font(name='PingFang TC', size=10, color='2980B9', italic=True)
    thin_border = Border(
        left=Side(style='thin', color='D5D8DC'),
        right=Side(style='thin', color='D5D8DC'),
        top=Side(style='thin', color='D5D8DC'),
        bottom=Side(style='thin', color='D5D8DC'),
    )
    even_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    
    # 欄寬
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 45
    
    # 標題
    ws.merge_cells('A1:B1')
    c = ws['A1']
    c.value = "PJSK 私車管理系統 — 指令手冊"
    c.font = title_font
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35
    
    row = 3
    for sec_name, cmds in sections:
        # 分類標題
        ws.merge_cells(f'A{row}:B{row}')
        c = ws[f'A{row}']
        c.value = f"▸ {sec_name}"
        c.font = section_font
        c.fill = section_fill
        c.alignment = Alignment(vertical='center')
        ws[f'B{row}'].fill = section_fill
        ws.row_dimensions[row].height = 28
        row += 1
        
        # 表頭
        for col, label in [('A','指令'), ('B','說明')]:
            c = ws[f'{col}{row}']
            c.value = label
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = thin_border
        ws.row_dimensions[row].height = 22
        row += 1
        
        # 指令列表
        for i, (cmd, desc) in enumerate(cmds):
            ca = ws[f'A{row}']
            cb = ws[f'B{row}']
            ca.value = cmd
            ca.font = cmd_font
            ca.border = thin_border
            ca.alignment = Alignment(vertical='center')
            cb.value = desc
            cb.font = desc_font
            cb.border = thin_border
            cb.alignment = Alignment(vertical='center')
            if i % 2 == 0:
                ca.fill = even_fill
                cb.fill = even_fill
            ws.row_dimensions[row].height = 22
            row += 1
        
        row += 1  # 分類間空行
    
    # 底部連結
    if link:
        ws.merge_cells(f'A{row}:B{row}')
        c = ws[f'A{row}']
        c.value = f"{link}"
        c.font = link_font
        c.alignment = Alignment(horizontal='center')
    
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    c = ws[f'A{row}']
    c.value = f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    c.font = Font(size=9, color='6C7A89')
    c.alignment = Alignment(horizontal='center')
    
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

@tree.command(name="help", description="指令手冊")
@app_commands.describe(模式="顯示模式")
@app_commands.choices(模式=[
    app_commands.Choice(name="文字（快速）", value="text"),
    app_commands.Choice(name="Excel 檔案", value="excel"),
    app_commands.Choice(name="圖片", value="image"),
])
async def help_cmd(interaction: discord.Interaction, 模式: str = "text"):
    sections=[
        ("成員管理 /成員",[("/成員 註冊","註冊（倍率、綜合力、多開、S6）"),("/成員 修改","修改個人資料"),
            ("/成員 查看","查看資料"),("/成員 列表","全體成員圖表"),("/成員 倍率計算","隊伍倍率公式")]),
        ("班表系統 /班表",[("/班表 開放 · 停止","管理員開/關報班"),("按鈕: 推手/S6/雙開/三開","快捷報班按鈕"),
            ("/班表 報班 · 取消","報名/取消時段"),("/班表 查看","今日班表圖片"),
            ("/班表 編輯 · 確認 · 清空","管理員操作"),("/班表 匯出 · 匯入","Excel 匯出/匯入"),
            ("快捷: /原推 · /s6 08-12","訊息快捷報班")]),
        ("房間管理 /房間",[("/房間 設定","房號+車種（選單）"),("/房間 換房 · 關閉","換房/關閉"),
            ("快捷: 設定房號 12345 蝦","訊息快捷設定")]),
        ("獎勵系統 /獎勵",[("/獎勵 發放","管理員發放 MyCard"),("/獎勵 查詢 · 統計","查詢/統計")]),
        ("查詢功能 /查詢",[("/查詢 活動排名","排名查詢（1h時速/上一局PT/場次）"),("/查詢 榜線","精彩片段榜線"),
            ("/查詢 榜線走勢 · 紀錄","走勢圖/紀錄"),("/查詢 玩家","查詢玩家個人檔案"),
            ("/查詢 控分","控分計算"),("/查詢 體力倍率","倍率對照表"),
            ("/查詢 時數","成員累計時數（原推/S6）"),("/查詢 時數匯出","匯出 Excel 時數表"),
            ("/查詢 個人時數","查看個人累計"),("快捷: e50 / e1-10","訊息快捷排名")]),
        ("實用工具 /工具",[("/工具 倒數","活動倒數計時"),("/工具 換算","分數萬位換算"),
            ("/工具 連結","PJSK 資源中心"),("/工具 肘人","肘人小幫手 (歌曲/火推薦)")]),
        ("系統管理 /系統（管理員）",[("/系統 時數歸零","清除所有歷史時數"),("/系統 備份","備份全部資料為 zip"),
            ("/系統 還原","從備份 zip 還原"),("/系統 狀態","查看系統狀態")]),
    ]
    if 模式 == "image":
        await interaction.response.defer()
        img=render_help_image("PJSK 私車管理系統",sections,link=PJSK_CENTER)
        await interaction.followup.send(file=discord.File(img,"help.png"),silent=True)
    elif 模式 == "excel":
        await interaction.response.defer()
        xlsx=render_help_excel(sections, link=PJSK_CENTER)
        await interaction.followup.send(file=discord.File(xlsx,"PJSK指令手冊.xlsx"),silent=True)
    else:
        lines = ["**PJSK 私車管理系統 — 指令手冊**\n"]
        for sec_name, cmds in sections:
            lines.append(f"**▸ {sec_name}**")
            for cmd, desc in cmds:
                lines.append(f"　`{cmd}` — {desc}")
            lines.append("")
        lines.append(f"{PJSK_CENTER}")
        await interaction.response.send_message("\n".join(lines),silent=True)

# ========== 成員指令 ==========
@grp_member.command(name="註冊", description="註冊資料")
@app_commands.describe(倍率="主帳倍率 (1.18~3.88)",綜合力="綜合力 (0~450000)",多開="多開",
    二開倍率="二開倍率",三開倍率="三開倍率",s6倍率="S6倍率",s6綜合="S6綜合",備註="備註")
@app_commands.choices(多開=[app_commands.Choice(name="單開",value="單開"),
    app_commands.Choice(name="雙開",value="雙開"),app_commands.Choice(name="三開",value="三開")])
async def register_cmd(interaction, 倍率:float, 綜合力:int, 多開:str="單開", 二開倍率:float=0.0,
                       三開倍率:float=0.0, s6倍率:float=0.0, s6綜合:int=0, 備註:str=""):
    uid=str(interaction.user.id)
    if not (1.18<=倍率<=3.88):
        await interaction.response.send_message("倍率範圍: 1.18~3.88",ephemeral=True,silent=True); return
    for nm,v in [("二開",二開倍率),("三開",三開倍率),("S6",s6倍率)]:
        if v and not (1.18<=v<=3.88):
            await interaction.response.send_message(f"{nm}倍率範圍: 1.18~3.88",ephemeral=True,silent=True); return
    bot_data["members"][uid]={"name":interaction.user.display_name,"bonus":float(倍率),"power":int(綜合力),
        "multi":多開,"bonus_2":float(二開倍率),"bonus_3":float(三開倍率),"s6_bonus":float(s6倍率),
        "s6_power":int(s6綜合),"note":備註,"registered_at":datetime.now().isoformat()}
    save_data()
    fields=[("名稱",interaction.user.display_name),("倍率",f"{倍率:.2f}"),("綜合力",f"{綜合力/10000:.2f}萬"),("多開",多開)]
    if 二開倍率>0: fields.append(("二開",f"{二開倍率:.2f}"))
    if 三開倍率>0: fields.append(("三開",f"{三開倍率:.2f}"))
    if s6倍率>0: fields.append(("S6倍率",f"{s6倍率:.2f}"))
    if s6綜合>0: fields.append(("S6綜合",f"{s6綜合/10000:.2f}萬"))
    img=render_info_card("註冊成功",fields,accent_color=Theme.GREEN)
    await interaction.response.send_message(file=discord.File(img,"reg.png"),silent=True)

@grp_member.command(name="修改", description="修改資料")
@app_commands.describe(倍率="主帳倍率",綜合力="綜合力",多開="多開",二開倍率="二開",三開倍率="三開",
    s6倍率="S6倍率",s6綜合="S6綜合",備註="備註")
async def update_cmd(interaction, 倍率:float=None, 綜合力:int=None, 多開:str=None, 二開倍率:float=None,
                     三開倍率:float=None, s6倍率:float=None, s6綜合:int=None, 備註:str=None):
    uid=str(interaction.user.id)
    if uid not in bot_data["members"]:
        img=render_message_box("錯誤",["請先 /成員 註冊"],accent_color=Theme.RED)
        await interaction.response.send_message(file=discord.File(img,"e.png"),silent=True); return
    m=bot_data["members"][uid]
    if 倍率 is not None: m["bonus"]=float(倍率)
    if 綜合力 is not None: m["power"]=int(綜合力)
    if 多開 is not None: m["multi"]=多開
    if 二開倍率 is not None: m["bonus_2"]=float(二開倍率)
    if 三開倍率 is not None: m["bonus_3"]=float(三開倍率)
    if s6倍率 is not None: m["s6_bonus"]=float(s6倍率)
    if s6綜合 is not None: m["s6_power"]=int(s6綜合)
    if 備註 is not None: m["note"]=備註
    save_data()
    img=render_info_card("已更新",[("倍率",f"{m.get('bonus',0):.2f}"),("綜合力",fmt_num(m.get('power',0))),
        ("多開",m.get('multi','單開'))],accent_color=Theme.GREEN)
    await interaction.response.send_message(file=discord.File(img,"u.png"),silent=True)

@grp_member.command(name="查看", description="查看資料")
async def my_cmd(interaction):
    uid=str(interaction.user.id)
    if uid not in bot_data["members"]:
        img=render_message_box("錯誤",["請先 /成員 註冊"],accent_color=Theme.RED)
        await interaction.response.send_message(file=discord.File(img,"e.png"),silent=True); return
    m=bot_data["members"][uid]
    fields=[("名稱",m.get('name','-')),("倍率",f"{m.get('bonus',0):.2f}"),
            ("綜合力",fmt_num(m.get('power',0))),("多開",m.get('multi','單開'))]
    if m.get('s6_bonus',0)>0: fields.append(("S6倍率",f"{m['s6_bonus']:.2f}"))
    if m.get('s6_power',0)>0: fields.append(("S6綜合",fmt_num(m['s6_power'])))
    img=render_info_card("個人資料",fields)
    await interaction.response.send_message(file=discord.File(img,"me.png"),silent=True)

@grp_member.command(name="列表", description="[管理員] 查看成員")
@admin_check()
async def member_list_cmd(interaction):
    await interaction.response.defer()
    members=bot_data.get("members",{})
    if not members:
        img=render_message_box("成員列表",["目前沒有成員"]); await interaction.followup.send(file=discord.File(img,"e.png"),silent=True); return
    img=create_member_table_image(members)
    await interaction.followup.send(file=discord.File(img,"members.png"),silent=True)

@grp_member.command(name="倍率計算", description="隊伍倍率計算")
@app_commands.describe(隊長倍率="隊長%",隊員1="隊員1%",隊員2="隊員2%",隊員3="隊員3%",隊員4="隊員4%")
async def calc_bonus_cmd(interaction, 隊長倍率:float, 隊員1:float, 隊員2:float, 隊員3:float, 隊員4:float):
    result=calculate_bonus(隊長倍率,[隊員1,隊員2,隊員3,隊員4])
    img=render_info_card("倍率計算",[("公式","[隊長%+100%+(隊員%總和/5)]/100%"),
        ("隊長",f"{隊長倍率}%"),("隊員總和",f"{隊員1+隊員2+隊員3+隊員4}%"),
        ("結果",f"{result:.2f}")],accent_color=Theme.BLUE)
    await interaction.response.send_message(file=discord.File(img,"calc.png"),silent=True)

# ========== 班表指令 ==========
@grp_schedule.command(name="開放", description="[管理員] 開放報班")
@admin_check()
async def open_cmd(interaction):
    bot_data.setdefault("settings",{})["schedule_open"]=True; save_data()
    img=render_message_box("報班系統",["報班已開放!","","使用下方按鈕或 /班表 報班"],accent_color=Theme.GREEN)
    await interaction.response.send_message(file=discord.File(img,"open.png"),view=ScheduleView(),silent=True)

@grp_schedule.command(name="停止", description="[管理員] 關閉報班")
@admin_check()
async def close_schedule_cmd(interaction):
    bot_data.setdefault("settings",{})["schedule_open"]=False; save_data()
    img=render_message_box("報班系統",["報班已關閉"],accent_color=Theme.RED)
    await interaction.response.send_message(file=discord.File(img,"close.png"),silent=True)

@grp_schedule.command(name="報班", description="報名時段")
@app_commands.describe(時段="如 08-12",角色="角色",備註="備註")
@app_commands.choices(角色=[app_commands.Choice(name="推手",value="pusher"),
    app_commands.Choice(name="S6",value="s6"),app_commands.Choice(name="外援",value="support")])
async def signup_cmd(interaction, 時段:str, 角色:str="pusher", 備註:str=""):
    uid=str(interaction.user.id)
    if uid not in bot_data.get("members",{}):
        img=render_message_box("錯誤",["請先 /成員 註冊"],accent_color=Theme.RED)
        await interaction.response.send_message(file=discord.File(img,"e.png"),silent=True); return
    if not bot_data.get("settings",{}).get("schedule_open"):
        img=render_message_box("錯誤",["報班未開放"],accent_color=Theme.RED)
        await interaction.response.send_message(file=discord.File(img,"e.png"),silent=True); return
    hours=parse_time_range(時段)
    if not hours:
        img=render_message_box("錯誤",["格式: 08-12"],accent_color=Theme.RED)
        await interaction.response.send_message(file=discord.File(img,"e.png"),silent=True); return
    today=get_today(); bot_data.setdefault("schedule",{}).setdefault(today,{})
    m=bot_data["members"][uid]
    app={"user_id":uid,"name":m["name"],"bonus":m["bonus"],"bonus_2":m.get("bonus_2",0),
         "bonus_3":m.get("bonus_3",0),"s6_bonus":m.get("s6_bonus",0),"power":m["power"],
         "s6_power":m.get("s6_power",0),"multi":m["multi"],"role":角色,"note":備註,
         "registered_at":datetime.now().isoformat()}
    registered=[]
    for h in hours:
        closed,reason=is_signup_closed(h)
        if closed: continue
        if h not in bot_data["schedule"][today]: bot_data["schedule"][today][h]={"applicants":[]}
        if not any(a["user_id"]==uid for a in bot_data["schedule"][today][h].get("applicants",[])):
            bot_data["schedule"][today][h].setdefault("applicants",[]).append(app); registered.append(h)
    save_data()
    if registered:
        refresh_schedule(today)
        img=render_message_box("報班成功",[f"時段: {', '.join(registered)}",f"角色: {角色}"],accent_color=Theme.GREEN)
        await interaction.response.send_message(file=discord.File(img,"ok.png"),silent=True)
    else:
        img=render_message_box("提示",["已報過或已截止"],accent_color=Theme.ORANGE)
        await interaction.response.send_message(file=discord.File(img,"dup.png"),silent=True)

@grp_schedule.command(name="取消", description="取消報班")
@app_commands.describe(時段="如 08-12")
async def cancel_cmd(interaction, 時段:str):
    uid=str(interaction.user.id); today=get_today(); hours=parse_time_range(時段); cancelled=[]
    for h in hours:
        if today in bot_data["schedule"] and h in bot_data["schedule"][today]:
            apps=bot_data["schedule"][today][h].get("applicants",[]); orig=len(apps)
            bot_data["schedule"][today][h]["applicants"]=[a for a in apps if a["user_id"]!=uid]
            if len(bot_data["schedule"][today][h]["applicants"])<orig: cancelled.append(h)
    save_data()
    if cancelled: refresh_schedule(today)
    msg="已取消: "+", ".join(cancelled) if cancelled else "無記錄"
    img=render_message_box("取消",[ msg],accent_color=Theme.GREEN if cancelled else Theme.ORANGE)
    await interaction.response.send_message(file=discord.File(img,"cancel.png"),silent=True)

@grp_schedule.command(name="查看", description="查看班表")
@app_commands.describe(模式="顯示模式")
@app_commands.choices(模式=[
    app_commands.Choice(name="圖片（預設）", value="image"),
    app_commands.Choice(name="Excel 檔案", value="excel"),
])
async def schedule_cmd(interaction, 模式: str = "image"):
    await interaction.response.defer()
    today=get_today(); schedule=bot_data.get("schedule",{}).get(today,{})
    if not schedule:
        await interaction.followup.send("今日無排班",silent=True); return
    if 模式 == "image":
        img=create_schedule_image(today,schedule)
        await interaction.followup.send(file=discord.File(img,"schedule.png"),view=ScheduleView(),silent=True)
    else:
        xlsx=create_schedule_excel(today, schedule)
        await interaction.followup.send(file=discord.File(xlsx,f"班表_{today}.xlsx"),view=ScheduleView(),silent=True)

@grp_schedule.command(name="編輯", description="[管理員] 手動編輯")
@admin_check()
async def edit_cmd(interaction): await interaction.response.send_modal(EditScheduleModal())

@grp_schedule.command(name="確認", description="[管理員] 確認排班")
@admin_check()
async def confirm_cmd(interaction):
    await interaction.response.defer(); today=get_today()
    if today not in bot_data.get("schedule",{}):
        await interaction.followup.send("無報班",silent=True); return
    for h in TIME_SLOTS:
        if h in bot_data["schedule"][today]:
            apps=bot_data["schedule"][today][h].get("applicants",[])
            bot_data["schedule"][today][h].update(auto_assign_schedule(today,h,apps))
    save_data()
    xlsx=create_schedule_excel(today,bot_data["schedule"][today])
    await interaction.followup.send("排班已確認",file=discord.File(xlsx,f"班表_{today}.xlsx"),view=ScheduleView(),silent=True)

@grp_schedule.command(name="清空", description="[管理員] 清空")
@admin_check()
async def clear_cmd(interaction):
    today=get_today()
    if today in bot_data.get("schedule",{}): del bot_data["schedule"][today]; save_data()
    img=render_message_box("已清空",[f"日期: {today}"],accent_color=Theme.RED)
    await interaction.response.send_message(file=discord.File(img,"clear.png"),silent=True)

@grp_schedule.command(name="匯出", description="匯出班表為 Excel")
@app_commands.describe(日期="日期 (留空為今天)")
async def export_csv_cmd(interaction, 日期:str=""):
    await interaction.response.defer()
    dt=日期.strip() if 日期.strip() else get_today()
    schedule=bot_data.get("schedule",{}).get(dt,{})
    if not schedule:
        await interaction.followup.send(f"{dt} 沒有班表資料",silent=True); return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "班表資料"
    
    # 樣式
    hdr_font = Font(name='PingFang TC', size=10, bold=True, color='FFFFFF')
    hdr_fill = PatternFill(start_color='1B2838', end_color='1B2838', fill_type='solid')
    data_font = Font(name='PingFang TC', size=10)
    thin_border = Border(
        left=Side(style='thin', color='D5D8DC'), right=Side(style='thin', color='D5D8DC'),
        top=Side(style='thin', color='D5D8DC'), bottom=Side(style='thin', color='D5D8DC'))
    center = Alignment(horizontal='center', vertical='center')
    even_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    active_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    
    # 標題列
    headers = ["時段","車種","P2(S6)","S6倍率","S6綜合","P3","P3倍率","P4","P4倍率","P5","P5倍率","外援","平均倍率","備註"]
    col_widths = [8, 8, 16, 10, 12, 16, 10, 16, 10, 16, 10, 16, 10, 14]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center; cell.border = thin_border
        ws.column_dimensions[chr(64+ci) if ci<=26 else 'A'+chr(64+ci-26)].width = w
    
    # 資料列
    members = bot_data.get("members", {})
    row_idx = 2
    for h in sorted(schedule.keys()):
        sh = schedule[h]
        if not isinstance(sh, dict): continue
        def gi(pos):
            p=sh.get(pos)
            if not p or not isinstance(p,dict): return ("","")
            b = p.get("bonus",0) or 0
            if b==0 and p.get("user_id") and p["user_id"] in members:
                b = members[p["user_id"]].get("bonus",0)
            return (p.get("name",""), f"{b:.2f}" if b > 0 else "")
        p2n,p2b=gi("p2"); p2d=sh.get("p2") or {}
        s6pw=p2d.get("s6_power",0) or p2d.get("power",0) or 0
        if s6pw==0 and p2d.get("user_id") and p2d["user_id"] in members:
            s6pw=members[p2d["user_id"]].get("s6_power",0) or members[p2d["user_id"]].get("power",0)
        p3n,p3b=gi("p3"); p4n,p4b=gi("p4"); p5n,p5b=gi("p5")
        sp=sh.get("support"); spn=sp.get("name","") if sp and isinstance(sp,dict) else ""
        
        vals = [h, sh.get("car_type","蝦"), p2n, p2b, str(s6pw) if s6pw>0 else "",
                p3n, p3b, p4n, p4b, p5n, p5b, spn, f"{sh.get('avg_bonus',0):.2f}", sh.get("note","")]
        
        has_data = p2n or p3n or p4n or p5n
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=ci, value=val)
            cell.font = data_font; cell.alignment = center; cell.border = thin_border
            if has_data:
                cell.fill = active_fill
            elif row_idx % 2 == 0:
                cell.fill = even_fill
        row_idx += 1
    
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    await interaction.followup.send(
        f"班表匯出完成 | 日期: {dt} | 共 {len(schedule)} 個時段",
        file=discord.File(buf, filename=f"班表_{dt}.xlsx"),silent=True)

@grp_schedule.command(name="匯入", description="[管理員] 從 Excel/CSV 匯入班表")
@admin_check()
@app_commands.describe(檔案="上傳 Excel 或 CSV",日期="日期 (留空為今天)")
async def import_csv_cmd(interaction, 檔案:discord.Attachment=None, 日期:str=""):
    await interaction.response.defer()
    dt=日期.strip() if 日期.strip() else get_today()
    attachment=檔案
    if not attachment:
        async for msg in interaction.channel.history(limit=10):
            for att in msg.attachments:
                if att.filename.endswith(('.csv','.xlsx','.xls')): attachment=att; break
            if attachment: break
    if not attachment:
        await interaction.followup.send(
            "**使用方式**\n"
            "方式1: `/班表 匯入 檔案:(拖入Excel或CSV)`\n"
            "方式2: 先上傳檔案到頻道再執行指令\n\n"
            "支援格式: `.xlsx` `.csv`",silent=True); return
    try:
        raw = await attachment.read()
    except Exception as e:
        await interaction.followup.send(f"無法讀取檔案: {e}",silent=True); return
    
    # 判斷格式並解析
    data_rows = []
    is_excel = attachment.filename.endswith(('.xlsx','.xls'))
    
    if is_excel:
        try:
            wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            all_rows = []
            for row in ws.iter_rows(values_only=True):
                # 正確處理各種 cell 值: None→"", 數字→保留精度, 其他→str
                cleaned = []
                for c in row:
                    if c is None:
                        cleaned.append("")
                    elif isinstance(c, float):
                        cleaned.append(f"{c:.2f}" if c != int(c) else str(int(c)))
                    else:
                        cleaned.append(str(c).strip())
                all_rows.append(cleaned)
            wb.close()
            if not all_rows:
                await interaction.followup.send("Excel 內容為空",silent=True); return
            # 判斷第一列是否為表頭
            header = all_rows[0]
            if header and not re.match(r'^\d', header[0].strip()):
                data_rows = all_rows[1:]
            else:
                data_rows = all_rows
        except Exception as e:
            await interaction.followup.send(f"Excel 解析失敗: {e}",silent=True); return
    else:
        try:
            try: csv_data = raw.decode('utf-8-sig')
            except: csv_data = raw.decode('utf-8')
        except Exception as e:
            await interaction.followup.send(f"無法讀取: {e}",silent=True); return
        try:
            dialect = csv.Sniffer().sniff(csv_data[:2000], delimiters=',;\t')
            reader = csv.reader(StringIO(csv_data), dialect)
        except:
            reader = csv.reader(StringIO(csv_data))
        rows = list(reader)
        if len(rows) < 2:
            await interaction.followup.send("CSV 內容為空",silent=True); return
        header = rows[0]
        if header and re.match(r'^\d', header[0].strip()):
            data_rows = rows
        else:
            data_rows = rows[1:]
    
    # 匯入邏輯（共用）
    bot_data.setdefault("schedule",{}).setdefault(dt,{})
    name_to_member={m.get("name",""):(uid,m) for uid,m in bot_data.get("members",{}).items()}
    imported=0
    for row in data_rows:
        if len(row)<2: continue
        h=row[0].strip()
        tm=re.match(r'^(\d{1,2}):?(\d{2})?',h)
        if not tm: continue
        hn=int(tm.group(1))
        if hn>23: continue
        h=f"{hn:02d}:00"
        ct=row[1].strip() if len(row)>1 else "蝦"
        shift={"car_type":ct if ct in CAR_TYPES else "蝦","p1":{"name":"omega","fixed":True},
               "p2":None,"p3":None,"p4":None,"p5":None,"support":None,"avg_bonus":0,"note":"","applicants":[]}
        def make_person(ni, bi, epi=None):
            name=row[ni].strip() if len(row)>ni else ""
            if not name: return None
            # 嘗試從 name 中提取倍率（如 "川風(2.50)"）
            name_match = re.match(r'^(.+?)\((\d+\.?\d*)\)$', name)
            extracted_bonus = 0.0
            if name_match:
                name = name_match.group(1)
                try: extracted_bonus = float(name_match.group(2))
                except: pass
            # 從倍率欄讀取
            bonus=0.0
            try:
                val = row[bi].strip() if len(row)>bi else ""
                if val: bonus=float(val)
            except: pass
            if bonus==0 and extracted_bonus>0: bonus=extracted_bonus
            person={"name":name,"bonus":bonus}
            if name in name_to_member:
                uid,m=name_to_member[name]; person["user_id"]=uid
                person.update({k:m.get(k,0) for k in ["power","s6_power","bonus_2","bonus_3","s6_bonus"]})
                person["multi"]=m.get("multi","單開")
                if bonus==0 or bonus<1.0: person["bonus"]=m.get("bonus",0)
            if epi is not None:
                try:
                    val = row[epi].strip() if len(row)>epi else ""
                    person["s6_power"]=int(float(val)) if val else 0
                except: pass
            return person
        p2=make_person(2,3,4)
        if p2: p2["role"]="s6"; shift["p2"]=p2
        p3=make_person(5,6)
        if p3: p3["role"]="pusher"; shift["p3"]=p3
        p4=make_person(7,8)
        if p4: p4["role"]="pusher"; shift["p4"]=p4
        p5=make_person(9,10)
        if p5: p5["role"]="pusher"; shift["p5"]=p5
        if len(row)>11 and row[11].strip():
            sn=row[11].strip(); support={"name":sn,"role":"support"}
            if sn in name_to_member: support["user_id"]=name_to_member[sn][0]
            shift["support"]=support
        if len(row)>13: shift["note"]=row[13].strip()
        bs=[shift[p].get('bonus',0) for p in ["p3","p4","p5"] if shift[p]]
        shift["avg_bonus"]=sum(bs)/len(bs) if bs else 0
        bot_data["schedule"][dt][h]=shift; imported+=1
    save_data()
    await interaction.followup.send(f"匯入完成 | 日期: {dt} | 匯入 {imported} 個時段",silent=True)

# ========== 成員累計時數系統 ==========
def count_member_hours():
    """統計所有成員的累計原推/S6時數"""
    stats = {}  # uid -> {"name":..., "pusher_hours":0, "s6_hours":0, "support_hours":0, "total_hours":0}
    for dt, schedule in bot_data.get("schedule",{}).items():
        for hour, shift in schedule.items():
            if not isinstance(shift, dict): continue
            # P3-P5 = 推手時數
            for pos in ["p3","p4","p5"]:
                p = shift.get(pos)
                if p and isinstance(p, dict) and p.get("user_id"):
                    uid = p["user_id"]
                    if uid not in stats:
                        stats[uid] = {"name":p.get("name","?"),"pusher_hours":0,"s6_hours":0,"support_hours":0}
                    stats[uid]["pusher_hours"] += 1
                    stats[uid]["name"] = p.get("name", stats[uid]["name"])
            # P2 = S6 時數
            p2 = shift.get("p2")
            if p2 and isinstance(p2, dict) and p2.get("user_id"):
                uid = p2["user_id"]
                if uid not in stats:
                    stats[uid] = {"name":p2.get("name","?"),"pusher_hours":0,"s6_hours":0,"support_hours":0}
                stats[uid]["s6_hours"] += 1
                stats[uid]["name"] = p2.get("name", stats[uid]["name"])
            # 外援
            sp = shift.get("support")
            if sp and isinstance(sp, dict) and sp.get("user_id"):
                uid = sp["user_id"]
                if uid not in stats:
                    stats[uid] = {"name":sp.get("name","?"),"pusher_hours":0,"s6_hours":0,"support_hours":0}
                stats[uid]["support_hours"] += 1
                stats[uid]["name"] = sp.get("name", stats[uid]["name"])
    # 合併成員表中有但班表中無紀錄的人
    for uid, m in bot_data.get("members",{}).items():
        if uid not in stats:
            stats[uid] = {"name":m.get("name","?"),"pusher_hours":0,"s6_hours":0,"support_hours":0}
    for uid in stats:
        s = stats[uid]
        s["total_hours"] = s["pusher_hours"] + s["s6_hours"] + s["support_hours"]
    return stats

def export_hours_excel(stats):
    """匯出累計時數為 Excel 檔案"""
    sorted_s = sorted(stats.items(), key=lambda x: x[1]["total_hours"], reverse=True)
    data = []
    for uid, s in sorted_s:
        m = bot_data.get("members",{}).get(uid,{})
        data.append({
            "名稱": s["name"],
            "倍率": m.get("bonus",0),
            "綜合力": m.get("power",0),
            "多開": m.get("multi","單開"),
            "原推時數": s["pusher_hours"],
            "S6時數": s["s6_hours"],
            "外援時數": s["support_hours"],
            "合計時數": s["total_hours"],
            "S6倍率": m.get("s6_bonus",0),
            "S6綜合": m.get("s6_power",0),
            "備註": m.get("note",""),
        })
    df = pd.DataFrame(data)
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='累計時數', index=False)
        # 設定欄寬
        ws = writer.sheets['累計時數']
        widths = [12,8,10,8,10,10,10,10,8,10,15]
        for i, w in enumerate(widths):
            col_letter = chr(65+i)
            ws.column_dimensions[col_letter].width = w
    buf.seek(0)
    return buf

@grp_query.command(name="時數", description="查看成員累計時數（圖片）")
async def hours_cmd(interaction: discord.Interaction):
    await interaction.response.defer()
    stats = count_member_hours()
    if not stats:
        img = render_message_box("時數統計",["尚無排班紀錄"])
        await interaction.followup.send(file=discord.File(img,"empty.png"),silent=True); return
    img = create_hours_table_image(stats)
    await interaction.followup.send(file=discord.File(img,"hours.png"),silent=True)

@grp_query.command(name="時數匯出", description="[管理員] 匯出累計時數為 Excel")
@admin_check()
async def hours_export_cmd(interaction: discord.Interaction):
    await interaction.response.defer()
    stats = count_member_hours()
    if not stats:
        img = render_message_box("時數統計",["尚無排班紀錄"])
        await interaction.followup.send(file=discord.File(img,"empty.png"),silent=True); return
    # 圖片
    img = create_hours_table_image(stats)
    await interaction.followup.send(file=discord.File(img,"hours.png"),silent=True)
    # Excel
    xlsx = export_hours_excel(stats)
    await interaction.followup.send(
        file=discord.File(xlsx, filename=f"member_hours_{get_today()}.xlsx"),silent=True)

@grp_query.command(name="個人時數", description="查看個人累計時數")
async def my_hours_cmd(interaction: discord.Interaction):
    uid = str(interaction.user.id)
    stats = count_member_hours()
    s = stats.get(uid)
    if not s:
        await interaction.response.send_message("尚無排班紀錄",silent=True); return
    # 排名
    sorted_s = sorted(stats.items(), key=lambda x: x[1]["total_hours"], reverse=True)
    rank = next((i+1 for i,(u,_) in enumerate(sorted_s) if u==uid), "-")
    msg = (
        f"**個人累計時數**\n\n"
        f"**名稱**: {s['name']}\n"
        f"**原推時數**: {s['pusher_hours']} h\n"
        f"**S6時數**: {s['s6_hours']} h\n"
        f"**外援時數**: {s['support_hours']} h\n"
        f"**合計**: {s['total_hours']} h\n"
        f"**排名**: 第 {rank} 名 / {len(stats)} 人"
    )
    await interaction.response.send_message(msg,silent=True)

# ========== 房間指令 ==========
@grp_room.command(name="設定", description="設定房間")
@app_commands.describe(房號="房號",車種="車種")
@app_commands.choices(車種=[app_commands.Choice(name=t,value=t) for t in CAR_TYPES])
async def room_cmd(interaction, 房號:str, 車種:str):
    ch=interaction.channel; orig=ch.name
    try: await ch.edit(name=f"{房號}-{車種}")
    except: pass
    bot_data["rooms"][str(ch.id)]={"room_id":房號,"car_type":車種,"original_name":orig,
        "created_at":datetime.now().isoformat(),"last_activity":datetime.now().isoformat()}
    save_data()
    img=render_info_card("房間設定",[("房號",房號),("車種",車種),("超時","30分鐘自動關閉")],accent_color=Theme.BLUE)
    await interaction.response.send_message(file=discord.File(img,"room.png"),silent=True)

@grp_room.command(name="換房", description="換房號")
@app_commands.describe(新房號="新房號")
async def change_cmd(interaction, 新房號:str):
    cid=str(interaction.channel_id)
    if cid not in bot_data["rooms"]:
        img=render_message_box("錯誤",["尚未設定"],accent_color=Theme.RED)
        await interaction.response.send_message(file=discord.File(img,"e.png"),silent=True); return
    ct=bot_data["rooms"][cid].get("car_type","蝦")
    try: await interaction.channel.edit(name=f"{新房號}-{ct}")
    except: pass
    bot_data["rooms"][cid]["room_id"]=新房號; bot_data["rooms"][cid]["last_activity"]=datetime.now().isoformat()
    save_data()
    img=render_message_box("換房",[f"新房號: {新房號}"],accent_color=Theme.GREEN)
    await interaction.response.send_message(file=discord.File(img,"ch.png"),silent=True)

@grp_room.command(name="關閉", description="關閉房間")
async def close_room_cmd(interaction):
    cid=str(interaction.channel_id); orig="私車"
    if cid in bot_data["rooms"]:
        orig=bot_data["rooms"][cid].get("original_name","私車"); del bot_data["rooms"][cid]; save_data()
    try: await interaction.channel.edit(name=orig)
    except: pass
    img=render_message_box("房間關閉",[f"頻道已恢復: {orig}"],accent_color=Theme.RED)
    await interaction.response.send_message(file=discord.File(img,"close.png"),silent=True)

# ========== 獎勵指令 ==========
@grp_reward.command(name="發放", description="[管理員] 發放 MyCard")
@admin_check()
@app_commands.describe(對象="成員",卡號="卡號",密碼="密碼",備註="備註")
async def issue_cmd(interaction, 對象:discord.Member, 卡號:str, 密碼:str, 備註:str=""):
    uid=str(對象.id)
    bot_data.setdefault("rewards",{}).setdefault(uid,[]).append({"card":卡號,"password":密碼,"note":備註,
        "issued_at":datetime.now().isoformat(),"issued_by":interaction.user.display_name})
    save_data()
    img=render_info_card("獎勵已發放",[("對象",對象.display_name)],accent_color=Theme.GREEN)
    await interaction.response.send_message(file=discord.File(img,"reward.png"),ephemeral=True,silent=True)
    try: await 對象.send("您收到了獎勵！使用 /獎勵 查詢 查看")
    except: pass

@grp_reward.command(name="查詢", description="查詢獎勵")
async def check_cmd(interaction):
    uid=str(interaction.user.id); rewards=bot_data.get("rewards",{}).get(uid,[])
    if not rewards:
        img=render_message_box("獎勵",["無獎勵"])
        await interaction.response.send_message(file=discord.File(img,"e.png"),ephemeral=True,silent=True); return
    fields=[]
    for i,r in enumerate(rewards,1):
        fields.append((f"獎勵{i} 卡號",r['card']))
        fields.append((f"獎勵{i} 密碼",r['password']))
    img=render_info_card(f"我的獎勵 ({len(rewards)}筆)",fields,accent_color=Theme.GOLD)
    await interaction.response.send_message(file=discord.File(img,"reward.png"),ephemeral=True,silent=True)

@grp_reward.command(name="統計", description="[管理員] 統計")
@admin_check()
async def reward_stats_cmd(interaction):
    rewards=bot_data.get("rewards",{}); total=sum(len(r) for r in rewards.values())
    fields=[("總發放",f"{total} 筆")]
    for uid,r in sorted(rewards.items(),key=lambda x:len(x[1]),reverse=True)[:5]:
        name=bot_data.get("members",{}).get(uid,{}).get("name",uid[:8])
        fields.append((name,f"{len(r)} 筆"))
    img=render_info_card("獎勵統計",fields,accent_color=Theme.GOLD)
    await interaction.response.send_message(file=discord.File(img,"stats.png"),ephemeral=True,silent=True)

# ========== 查詢指令 ==========
@grp_query.command(name="體力倍率", description="體力倍率對照表")
async def energy_cmd(interaction):
    headers=["消耗體力","倍率"]
    rows=[[str(e),f"{m}x"] for e,m in ENERGY_MULTIPLIERS.items()]
    img=render_table_image(title="體力倍率表",subtitle="消耗體力 → 分數倍率",
        headers=headers,rows=rows,col_widths=[0.5,0.5],
        col_colors={1:Theme.RED},figsize=(6,8))
    await interaction.response.send_message(file=discord.File(img,"energy.png"),silent=True)

@grp_query.command(name="活動排名", description="查詢活動排名")
@app_commands.describe(名次="指定名次 (留空前10)")
async def ranking_cmd(interaction, 名次:int=0):
    await interaction.response.defer()
    try:
        async with ClientSession() as session:
            async with session.get(f"{HISEKAI_API}/event/live/top100",timeout=ClientTimeout(total=15)) as resp:
                data=await resp.json()
        rankings=data.get('top_100_player_rankings',[]); event_name=data.get('name','-')
        if 名次>0:
            target=None; prev_p=None; next_p=None
            for p in rankings:
                if p.get('rank')==名次: target=p
                if p.get('rank')==名次-1: prev_p=p
                if p.get('rank')==名次+1: next_p=p
            if not target: await interaction.followup.send(f"找不到第{名次}名",silent=True); return
            # 歷史走勢
            rk=str(名次); history_data=[]
            cur_recs=[r for r in ranking_history.get("records",[]) if r.get('event')==event_name]
            for rec in cur_recs:
                if rk in rec.get("borders",{}): history_data.append({'time':rec['time'],'score':rec["borders"][rk]["score"]})
            img=create_ranking_detail_image(target,prev_p,next_p,event_name,history_data)
            if img: await interaction.followup.send(file=discord.File(img,f"rank{名次}.png"),silent=True)
        else:
            img=create_ranking_list_image(rankings,1,10,event_name)
            if img: await interaction.followup.send(file=discord.File(img,"top10.png"),view=RankQueryView(),silent=True)
            else: await interaction.followup.send("無法生成",silent=True)
    except Exception as e: await interaction.followup.send(f"查詢失敗: {e}",silent=True)

class RankQueryView(View):
    def __init__(self): super().__init__(timeout=120)
    @discord.ui.button(label="T1",style=discord.ButtonStyle.danger,row=0)
    async def t1(self,i,b): await self._q(i,1)
    @discord.ui.button(label="T2",style=discord.ButtonStyle.primary,row=0)
    async def t2(self,i,b): await self._q(i,2)
    @discord.ui.button(label="T3",style=discord.ButtonStyle.primary,row=0)
    async def t3(self,i,b): await self._q(i,3)
    @discord.ui.button(label="T10",style=discord.ButtonStyle.secondary,row=0)
    async def t10(self,i,b): await self._q(i,10)
    @discord.ui.button(label="T50",style=discord.ButtonStyle.secondary,row=1)
    async def t50(self,i,b): await self._q(i,50)
    @discord.ui.button(label="T100",style=discord.ButtonStyle.secondary,row=1)
    async def t100(self,i,b): await self._q(i,100)
    @discord.ui.button(label="走勢圖",style=discord.ButtonStyle.success,emoji="📈",row=1)
    async def chart(self,interaction,button):
        await interaction.response.defer()
        img=create_ranking_chart()
        if img: await interaction.followup.send(file=discord.File(img,"chart.png"),silent=True)
        else: await interaction.followup.send("紀錄不足",ephemeral=True,silent=True)
    async def _q(self,interaction,rank):
        await interaction.response.defer()
        try:
            async with ClientSession() as s:
                async with s.get(f"{HISEKAI_API}/event/live/top100",timeout=ClientTimeout(total=15)) as r:
                    data=await r.json()
            rankings=data.get('top_100_player_rankings',[]); ev=data.get('name','-')
            target=prev_p=next_p=None
            for p in rankings:
                if p.get('rank')==rank: target=p
                if p.get('rank')==rank-1: prev_p=p
                if p.get('rank')==rank+1: next_p=p
            if not target: await interaction.followup.send(f"找不到T{rank}",ephemeral=True,silent=True); return
            rk=str(rank); hd_list=[]
            cur=[r for r in ranking_history.get("records",[]) if r.get('event')==ev]
            for rec in cur:
                if rk in rec.get("borders",{}): hd_list.append({'time':rec['time'],'score':rec["borders"][rk]["score"]})
            img=create_ranking_detail_image(target,prev_p,next_p,ev,hd_list)
            if img: await interaction.followup.send(file=discord.File(img,f"t{rank}.png"),silent=True)
        except Exception as e: await interaction.followup.send(f"錯誤: {e}",ephemeral=True,silent=True)

@grp_query.command(name="榜線走勢", description="榜線走勢圖")
@app_commands.describe(名次="指定名次 (留空全部)")
async def ranking_chart_cmd(interaction, 名次:int=0):
    await interaction.response.defer()
    img=create_ranking_chart(名次 if 名次>0 else None)
    if img: await interaction.followup.send(file=discord.File(img,"chart.png"),silent=True)
    else: await interaction.followup.send("紀錄不足 (需≥2筆)",silent=True)

@grp_query.command(name="榜線", description="查詢精彩片段榜線")
async def border_cmd(interaction):
    await interaction.response.defer()
    try:
        async with ClientSession() as session:
            async with session.get(f"{HISEKAI_API}/event/live/border",timeout=ClientTimeout(total=15)) as resp:
                data=await resp.json()
        borders=data.get('border_player_rankings',[]); event_name=data.get('name','-')
        if not borders: await interaction.followup.send("目前無榜線資料",silent=True); return
        headers=["排名","玩家名稱","總分","上一局PT","1h時速","場次(1h)"]
        rows=[]
        for p in borders:
            rk=p.get('rank',0); sc=p.get('score',0)
            h1 = p.get('last_1h_stats') or {}
            last_sc = p.get('last_score', 0)
            last_pt = f"{last_sc/10000:.4f}W" if last_sc else "-"
            speed_1h = f"{h1['speed']/10000:.2f}W/h" if h1.get('speed') else "-"
            count_1h = str(h1.get('count', 0)) if h1.get('count') else "-"
            rows.append([f"#{rk}", p.get('name','-'), f"{sc/10000:,.4f}W", last_pt, speed_1h, count_1h])
        rh={i:'#E8D5A8' for i,p in enumerate(borders) if p.get('rank',999)<=3}
        img=render_table_image(title="精彩片段榜線", subtitle=event_name,
            headers=headers, rows=rows, col_widths=[0.08,0.24,0.18,0.18,0.16,0.10],
            col_colors={0:Theme.RED,2:Theme.BLUE,3:Theme.PURPLE,4:Theme.GREEN}, row_highlights=rh,
            footer=f"更新: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | 資料來源: hisekai.org",
            figsize=(16, max(6, 2.5+len(rows)*0.55)))
        if img: await interaction.followup.send(file=discord.File(img,"border.png"),silent=True)
    except Exception as e: await interaction.followup.send(f"查詢失敗: {e}",silent=True)

@grp_query.command(name="玩家", description="查詢玩家個人檔案")
@app_commands.describe(玩家id="玩家 ID")
async def player_profile_cmd(interaction, 玩家id: str):
    await interaction.response.defer()
    try:
        async with ClientSession() as session:
            async with session.get(f"{HISEKAI_API}/user/{玩家id.strip()}/profile",timeout=ClientTimeout(total=15)) as resp:
                if resp.status != 200:
                    await interaction.followup.send(f"查詢失敗 (HTTP {resp.status})",silent=True); return
                data=await resp.json()
        # 基本資料
        user = data.get('user', data)  # 嘗試取 user 或直接用 data
        uid = user.get('userId', user.get('id', 玩家id))
        name = user.get('name', '-')
        rank_val = user.get('rank', '-')
        word = user.get('word', user.get('profile', {}).get('word', '-')) or '-'
        twitter = user.get('twitterId', user.get('profile', {}).get('twitter_id', '')) or '-'
        
        # 嘗試讀取更多資料
        user_decks = data.get('userDecks', [])
        user_cards = data.get('userCards', [])
        total_power = '-'
        if user_decks:
            # 取第一組隊伍的綜合力
            deck = user_decks[0] if user_decks else {}
            tp = deck.get('totalPower', 0)
            if tp: total_power = f"{tp/10000:.2f}萬" if tp>=10000 else str(tp)
        
        challenge_live_rank = data.get('userChallengeLiveSoloResult', {}).get('highScore', '-')
        
        msg = (
            f"**玩家檔案**\n\n"
            f"**名稱**: {name}\n"
            f"**ID**: {uid}\n"
            f"**等級**: {rank_val}\n"
            f"**綜合力**: {total_power}\n"
            f"**簽名**: {word}\n"
            f"**Twitter**: {twitter}\n"
            f"**卡片數**: {len(user_cards)} 張\n"
        )
        await interaction.followup.send(msg,silent=True)
    except Exception as e: await interaction.followup.send(f"查詢失敗: {e}",silent=True)

@grp_query.command(name="控分", description="控分計算")
@app_commands.describe(目標分數="目標",目前分數="目前")
async def score_cmd(interaction, 目標分數:int, 目前分數:int):
    if not table:
        img=render_message_box("錯誤",["分數表未載入"],accent_color=Theme.RED)
        await interaction.response.send_message(file=discord.File(img,"e.png"),silent=True); return
    diff=目標分數-目前分數
    if diff<=0:
        img=render_message_box("錯誤",["目標須大於目前"],accent_color=Theme.RED)
        await interaction.response.send_message(file=discord.File(img,"e.png"),silent=True); return
    solution=find_solution(table,diff)
    fields=[("目標",f"{目標分數:,}"),("目前",f"{目前分數:,}"),("差分",f"{diff:,}")]
    if not solution: fields.append(("結果","找不到組合"))
    else:
        for i,s in enumerate(solution,1):
            fields.append((f"Step{i}",f"{s['range']} | {int(s['bonus']*100)}% | 體{s['energy']} | {s['plays']}次 | +{s['total']:,}"))
    img=render_info_card("控分系統",fields,accent_color=Theme.BLUE)
    await interaction.response.send_message(file=discord.File(img,"score.png"),silent=True)

@grp_query.command(name="統計", description="[管理員] 統計資料")
@admin_check()
async def stats_cmd(interaction):
    members=bot_data.get("members",{}); rewards=bot_data.get("rewards",{})
    bonuses=[m.get("bonus",0) for m in members.values()]
    avg=sum(bonuses)/len(bonuses) if bonuses else 0
    multi={"單開":0,"雙開":0,"三開":0}
    for m in members.values(): multi[m.get("multi","單開")]=multi.get(m.get("multi","單開"),0)+1
    fields=[("成員數",str(len(members))),("平均倍率",f"{avg:.2f}"),
            ("單開",f"{multi['單開']}人"),("雙開",f"{multi['雙開']}人"),("三開",f"{multi['三開']}人"),
            ("獎勵",f"{sum(len(r) for r in rewards.values())}筆")]
    img=render_info_card("統計",fields)
    await interaction.response.send_message(file=discord.File(img,"stats.png"),ephemeral=True,silent=True)

# ========== 工具指令 ==========
@grp_tools.command(name="倒數", description="活動倒數計時")
@app_commands.describe(結束時間="活動結束時間 (格式: 2026-02-15 20:00)")
async def countdown_cmd(interaction, 結束時間:str):
    try:
        end=datetime.strptime(結束時間,"%Y-%m-%d %H:%M"); now=datetime.now()
        diff=end-now
        if diff.total_seconds()<=0:
            img=render_message_box("倒數",["活動已結束！"],accent_color=Theme.RED)
        else:
            d=diff.days; h,rem=divmod(diff.seconds,3600); m,s=divmod(rem,60)
            img=render_info_card("活動倒數",[("剩餘時間",f"{d}天 {h}時 {m}分"),
                ("結束時間",結束時間)],accent_color=Theme.PINK)
        await interaction.response.send_message(file=discord.File(img,"cd.png"),silent=True)
    except:
        img=render_message_box("錯誤",["格式: 2026-02-15 20:00"],accent_color=Theme.RED)
        await interaction.response.send_message(file=discord.File(img,"e.png"),silent=True)

@grp_tools.command(name="換算", description="分數萬位換算")
@app_commands.describe(分數="輸入分數")
async def convert_cmd(interaction, 分數:int):
    img=render_info_card("分數換算",[("原始",f"{分數:,}"),("萬位",f"{分數/10000:.4f}W"),
        ("億位",f"{分數/100000000:.8f}億")],accent_color=Theme.BLUE)
    await interaction.response.send_message(file=discord.File(img,"conv.png"),silent=True)

@grp_tools.command(name="連結", description="PJSK 資源中心")
async def link_cmd(interaction):
    msg = (
        "**PJSK 資源中心**\n\n"
        f"**網站**: {PJSK_CENTER}\n"
        "**功能**: 活動資訊、卡片查詢、音樂列表\n"
        "**API**: hisekai.org（排名資料來源）"
    )
    await interaction.response.send_message(msg,silent=True)

@grp_tools.command(name="肘人", description="肘人小幫手 — 找出最佳歌曲/火力方案追上指定名次")
@app_commands.describe(
    目標名次="想肘到的名次 (1~100)",
    目前ep="你目前的 EP",
    綜合力="隊伍綜合力",
    加成="活動加成百分比 (如 250)",
    倍率="技能倍率 (3.2 = 實效 2.2，預設 3.2)",
    s6倍率="S6 技能倍率 (3.2 = 實效 2.2，預設 3.2)",
    間隔秒數="歌曲間等待秒數 (預設 50)"
)
async def push_cmd(interaction, 目標名次:int, 目前ep:int, 綜合力:int, 加成:int=250,
                   倍率:float=3.2, s6倍率:float=3.2, 間隔秒數:int=50):
    await interaction.response.defer()
    
    if not SONG_DB:
        await interaction.followup.send("歌曲資料庫尚未載入，請聯繫管理員。",silent=True)
        return
    
    if 目標名次 < 1 or 目標名次 > 100:
        await interaction.followup.send("名次範圍為 1~100。",silent=True)
        return
    
    # 倍率轉實效 (3.2 → 實效 2.2)
    effective_skill = 倍率 - 1.0
    effective_s6 = s6倍率 - 1.0
    
    try:
        target_score = 0
        event_name = ""
        border_info = {'name': '???', 'speed_1h': 0, 'speed_3h': 0, 'speed_24h': 0}
        
        async with ClientSession() as session:
            # top100 — 取分數 + 榜線速度資訊
            async with session.get(f"{HISEKAI_API}/event/live/top100", timeout=ClientTimeout(total=15)) as resp:
                if resp.status == 200:
                    data = await resp.json()
                    rankings = data.get('top_100_player_rankings', [])
                    event_name = data.get('name', '')
                    for p in rankings:
                        if p.get('rank') == 目標名次:
                            target_score = p.get('score', 0)
                            border_info['name'] = p.get('name', '???')
                            h1 = p.get('last_1h_stats') or {}
                            h3 = p.get('last_3h_stats') or {}
                            h24 = p.get('last_24h_stats') or {}
                            border_info['speed_1h'] = h1.get('speed', 0)
                            border_info['speed_3h'] = h3.get('speed', 0)
                            border_info['speed_24h'] = h24.get('speed', 0)
                            border_info['last_played_at'] = p.get('last_played_at', '')
                            break
            
            # 如果 top100 找不到，試 border
            if target_score == 0:
                async with session.get(f"{HISEKAI_API}/event/live/border", timeout=ClientTimeout(total=15)) as resp:
                    if resp.status == 200:
                        data = await resp.json()
                        borders = data.get('border_player_rankings', [])
                        if not event_name:
                            event_name = data.get('name', '')
                        for p in borders:
                            if p.get('rank') == 目標名次:
                                target_score = p.get('score', 0)
                                border_info['name'] = p.get('name', '???')
                                h1 = p.get('last_1h_stats') or {}
                                h3 = p.get('last_3h_stats') or {}
                                h24 = p.get('last_24h_stats') or {}
                                border_info['speed_1h'] = h1.get('speed', 0)
                                border_info['speed_3h'] = h3.get('speed', 0)
                                border_info['speed_24h'] = h24.get('speed', 0)
                                border_info['last_played_at'] = p.get('last_played_at', '')
                                break
        
        if target_score == 0:
            await interaction.followup.send(f"無法取得第 {目標名次} 名的分數，可能排名資料尚未更新。",silent=True)
            return
        
        gap = target_score - 目前ep
        if gap <= 0:
            await interaction.followup.send(
                f"你的 EP ({目前ep/10000:,.2f}W) 已經超過第 {目標名次} 名 ({target_score/10000:,.2f}W)，不需要肘人！",silent=True)
            return
        
        # 取榜線速度 (優先 1h > 3h > 24h)
        border_speed = border_info.get('speed_1h') or border_info.get('speed_3h') or border_info.get('speed_24h') or 0
        
        # 找方案
        plans = find_push_plans(
            target_ep_gap=gap,
            power=綜合力,
            bonus=加成,
            skill_mag=effective_skill,
            s6=effective_s6,
            live_type='multi',
            interval=間隔秒數,
            energy_options=[5, 7, 10],
            border_speed=border_speed
        )
        
        if not plans:
            await interaction.followup.send("找不到可行方案（你的時速可能追不上榜線速度），請確認參數。",silent=True)
            return
        
        # 按體力分組 (收集足夠方案供兩種排序)
        grouped_txt = {}
        for p in plans:
            e = p['energy']
            if e not in grouped_txt:
                grouped_txt[e] = []
            if len(grouped_txt[e]) < 15:
                grouped_txt[e].append(p)
        
        # 文字版 (手機友善，每段top3)
        def fmt_row(ri, r):
            ap = r.get('adj_plays', r['plays'])
            at = r.get('adj_time_min', r['time_min'])
            ast_ = r.get('adj_stamina', r['stamina'])
            t_str = f"{at/60:.1f}h" if at >= 60 else f"{at:.0f}m"
            return f"{ri+1}.{r['title'][:8]} {r['diff']}{r['lv']} {r['eph']:,}/h {ap}場 {t_str} {ast_}體"
        
        tl = []
        tl.append(f"─ 肘人 No.{目標名次} {border_info.get('name','???')[:8]} {target_score/10000:,.2f}W")
        tl.append(f"目前{目前ep/10000:,.2f}W 差{gap/10000:,.2f}W 力{綜合力:,} 加{加成}%")
        if border_speed > 0:
            bs_label = "1h" if border_info.get('speed_1h') else ("3h" if border_info.get('speed_3h') else "24h")
            tl.append(f"榜線{border_speed/10000:,.4f}W/h({bs_label})")
        
        tl.append("【長效】EP效率")
        for energy in sorted(grouped_txt.keys()):
            rows = sorted(grouped_txt[energy], key=lambda x: -x['eph'])[:3]
            tl.append(f"▸x{energy}火")
            for ri, r in enumerate(rows):
                tl.append(fmt_row(ri, r))
        
        tl.append("【短效】最快")
        for energy in sorted(grouped_txt.keys()):
            rows = sorted(grouped_txt[energy], key=lambda x: x.get('adj_plays', x['plays']))[:3]
            tl.append(f"▸x{energy}火")
            for ri, r in enumerate(rows):
                tl.append(fmt_row(ri, r))
        
        if border_speed > 0:
            tl.append("*含榜線追趕修正")
        
        text_msg = "```\n" + "\n".join(tl) + "\n```"
        # Discord 2000 char limit - split if needed
        if len(text_msg) > 1950:
            # 拆成長效+短效兩條
            split_idx = "\n".join(tl).find("【短效】")
            if split_idx > 0:
                part1 = "```\n" + "\n".join(tl)[:split_idx].rstrip() + "\n```"
                part2 = "```\n" + "\n".join(tl)[split_idx:] + "\n```"
                if len(part1) > 1950: part1 = part1[:1947] + "```"
                if len(part2) > 1950: part2 = part2[:1947] + "```"
                text_msg = part1
                text_msg_2 = part2
            else:
                text_msg = text_msg[:1947] + "```"
                text_msg_2 = None
        else:
            text_msg_2 = None
        
        # 渲染圖片 (嘗試遠端 → 降級本地)
        render_kwargs = dict(plans=plans, target_rank=目標名次, target_score=target_score,
                            current_ep=目前ep, gap=gap, power=綜合力, bonus=加成,
                            event_name=event_name, border_info=border_info)
        buf = None
        try:
            buf = await _remote_render('create_push_plan_image', **render_kwargs)
            if buf is None:
                buf = create_push_plan_image(**render_kwargs)
        except Exception as img_err:
            import traceback
            print(f"[push_cmd] Image render error: {traceback.format_exc()}")
        
        if buf:
            await interaction.followup.send(content=text_msg, file=discord.File(buf, "push_plan.png"),silent=True)
        else:
            await interaction.followup.send(text_msg,silent=True)
        if text_msg_2:
            await interaction.followup.send(text_msg_2,silent=True)
    
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"[push_cmd] Error: {tb}")
        err_msg = f"錯誤: {str(e)[:500]}"
        await interaction.followup.send(err_msg,silent=True)

# ========== 系統管理 /系統 ==========
@grp_system.command(name="時數歸零", description="[管理員] 歸零全員累計時數")
@admin_check()
@app_commands.describe(確認="輸入 CONFIRM 確認歸零")
async def reset_hours_cmd(interaction: discord.Interaction, 確認: str = ""):
    if 確認 != "CONFIRM":
        await interaction.response.send_message(
            "**此操作會清除所有排班歷史紀錄（時數歸零）**\n"
            "今日班表會保留，僅清除過去資料。\n\n"
            "確認請輸入: `/系統 時數歸零 確認:CONFIRM`", ephemeral=True,silent=True); return
    await interaction.response.defer()
    today = get_today()
    old_schedule = bot_data.get("schedule", {})
    today_data = old_schedule.get(today, {})
    cleared = len(old_schedule) - (1 if today_data else 0)
    # 只保留今日
    bot_data["schedule"] = {today: today_data} if today_data else {}
    save_data()
    await interaction.followup.send(
        f"**時數已歸零**\n"
        f"清除 {cleared} 天的排班紀錄\n"
        f"今日班表已保留",silent=True)

@grp_system.command(name="備份", description="[管理員] 備份系統資料")
@admin_check()
async def backup_cmd(interaction: discord.Interaction):
    await interaction.response.defer()
    now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    buf = BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        # 主資料
        if os.path.exists(DATA_FILE):
            zf.write(DATA_FILE, DATA_FILE)
        # 排名歷史
        if os.path.exists(RANKING_HISTORY_FILE):
            zf.write(RANKING_HISTORY_FILE, RANKING_HISTORY_FILE)
        # 寫入備份資訊
        info = json.dumps({
            "backup_time": datetime.now().isoformat(),
            "members": len(bot_data.get("members", {})),
            "schedule_days": len(bot_data.get("schedule", {})),
            "rewards": len(bot_data.get("rewards", {})),
        }, ensure_ascii=False, indent=2)
        zf.writestr("backup_info.json", info)
    buf.seek(0)
    
    stats = count_member_hours()
    total_hours = sum(s["total_hours"] for s in stats.values())
    await interaction.followup.send(
        f"**系統備份完成** — {now_str}\n"
        f"成員: {len(bot_data.get('members',{}))} 人 | "
        f"班表: {len(bot_data.get('schedule',{}))} 天 | "
        f"累計時數: {total_hours} h",
        file=discord.File(buf, f"pjsk_backup_{now_str}.zip"),silent=True)

@grp_system.command(name="還原", description="[管理員] 從備份還原")
@admin_check()
@app_commands.describe(檔案="上傳備份 zip 檔")
async def restore_cmd(interaction: discord.Interaction, 檔案: discord.Attachment = None):
    await interaction.response.defer()
    attachment = 檔案
    if not attachment:
        async for msg in interaction.channel.history(limit=10):
            for att in msg.attachments:
                if att.filename.endswith('.zip'): attachment = att; break
            if attachment: break
    if not attachment:
        await interaction.followup.send(
            "**使用方式**\n"
            "`/系統 還原 檔案:(拖入備份zip)`\n"
            "或先上傳 zip 到頻道再執行指令", ephemeral=True,silent=True); return
    try:
        raw = await attachment.read()
        zf = zipfile.ZipFile(BytesIO(raw), 'r')
        names = zf.namelist()
        if DATA_FILE not in names:
            await interaction.followup.send("無效的備份檔（找不到資料檔）", ephemeral=True,silent=True); return
        # 先備份當前
        now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        if os.path.exists(DATA_FILE):
            shutil.copy2(DATA_FILE, f"{DATA_FILE}.before_restore_{now_str}")
        # 還原
        restored = []
        for name in [DATA_FILE, RANKING_HISTORY_FILE]:
            if name in names:
                with open(name, 'wb') as f:
                    f.write(zf.read(name))
                restored.append(name)
        zf.close()
        # 重新載入
        global bot_data, ranking_history
        bot_data = load_json(DATA_FILE, bot_data)
        ranking_history = load_json(RANKING_HISTORY_FILE, ranking_history)
        
        await interaction.followup.send(
            f"**還原完成**\n"
            f"還原檔案: {', '.join(restored)}\n"
            f"成員: {len(bot_data.get('members',{}))} 人 | "
            f"班表: {len(bot_data.get('schedule',{}))} 天\n"
            f"還原前的備份已儲存為 `{DATA_FILE}.before_restore_{now_str}`",silent=True)
    except Exception as e:
        await interaction.followup.send(f"還原失敗: {e}", ephemeral=True,silent=True)

@grp_system.command(name="狀態", description="[管理員] 查看系統狀態")
@admin_check()
async def status_cmd(interaction: discord.Interaction):
    stats = count_member_hours()
    total_hours = sum(s["total_hours"] for s in stats.values())
    schedule_days = len(bot_data.get("schedule", {}))
    rooms = len(bot_data.get("rooms", {}))
    members = len(bot_data.get("members", {}))
    sched_open = "開放" if bot_data.get("settings",{}).get("schedule_open") else "關閉"
    
    # 資料檔大小
    data_size = os.path.getsize(DATA_FILE) / 1024 if os.path.exists(DATA_FILE) else 0
    
    msg = (
        f"**系統狀態**\n\n"
        f"**成員**: {members} 人\n"
        f"**班表**: {schedule_days} 天紀錄\n"
        f"**累計時數**: {total_hours} h\n"
        f"**活躍房間**: {rooms} 間\n"
        f"**報班狀態**: {sched_open}\n"
        f"**資料大小**: {data_size:.1f} KB\n"
        f"**運行時間**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    )
    await interaction.response.send_message(msg, ephemeral=True,silent=True)

# ========== 註冊指令群組 ==========
for g in [grp_member,grp_schedule,grp_room,grp_reward,grp_query,grp_tools,grp_system]:
    tree.add_command(g)

# ========== 背景任務 ==========
async def reminder_task():
    await client.wait_until_ready()
    while not client.is_closed():
        try:
            now=datetime.now()
            # 50分提醒
            if now.minute==50:
                today=get_today(); next_hour=f"{(now.hour+1)%24:02d}:00"
                if today in bot_data.get("schedule",{}):
                    shift=bot_data["schedule"][today].get(next_hour,{})
                    if shift:
                        mentions=[]
                        for pos in ["p2","p3","p4","p5","support"]:
                            if shift.get(pos) and shift[pos].get('user_id'):
                                m=f"<@{shift[pos]['user_id']}>"
                                if m not in mentions: mentions.append(m)
                        if mentions:
                            for guild in client.guilds:
                                ch=discord.utils.get(guild.text_channels,name="排班提醒")
                                if not ch: ch=discord.utils.get(guild.text_channels,name="私車")
                                if ch:
                                    img=render_message_box("排班提醒",[f"時段: {next_hour}",
                                        f"車種: {shift.get('car_type','蝦')}",
                                        f"平均倍率: {shift.get('avg_bonus',0):.2f}","","請準備上車!"],accent_color=Theme.ORANGE)
                                    await ch.send(" ".join(mentions),file=discord.File(img,"remind.png"),silent=True)
            # 房間超時
            for cid in list(bot_data.get("rooms",{}).keys()):
                info=bot_data["rooms"][cid]
                last=datetime.fromisoformat(info.get("last_activity",datetime.now().isoformat()))
                if (datetime.now()-last).total_seconds()>1800:
                    try:
                        ch=client.get_channel(int(cid))
                        if ch:
                            orig=info.get("original_name","私車"); await ch.edit(name=orig)
                            img=render_message_box("房間關閉",["30分鐘無活動",f"已恢復: {orig}"],accent_color=Theme.RED)
                            await ch.send(file=discord.File(img,"timeout.png"),silent=True)
                        del bot_data["rooms"][cid]; save_data()
                    except: pass
            # 每小時記錄排名
            if now.minute<=1:
                try: await record_ranking_snapshot()
                except Exception as e: print(f"[Ranking Error] {e}")
        except Exception as e: print(f"[Task Error] {e}")
        await asyncio.sleep(60)

async def record_ranking_snapshot():
    global ranking_history
    async with ClientSession() as session:
        async with session.get(f"{HISEKAI_API}/event/live/top100",timeout=ClientTimeout(total=15)) as resp:
            if resp.status!=200: return
            data=await resp.json()
    rankings=data.get('top_100_player_rankings',[]); event_name=data.get('name','')
    if not rankings: return
    ranking_history["event_name"]=event_name
    now_str=datetime.now().strftime("%Y-%m-%d %H:00")
    records=ranking_history.setdefault("records",[])
    if records and records[-1].get("time","").startswith(now_str[:13]): return
    snapshot={"time":now_str,"event":event_name,"borders":{}}
    for p in rankings:
        r=p.get('rank')
        if r: snapshot["borders"][str(r)]={"name":p.get('name','-'),"score":p.get('score',0)}
    records.append(snapshot)
    if len(records)>336: ranking_history["records"]=records[-336:]
    save_ranking()

# ========== on_message 快捷指令 ==========
@tree.error
async def on_app_command_error(interaction: discord.Interaction, error):
    if isinstance(error, app_commands.CheckFailure):
        # admin_check 已經回覆了，忽略
        if not interaction.response.is_done():
            await interaction.response.send_message("權限不足", ephemeral=True,silent=True)
        return
    # 其他錯誤正常拋出
    raise error

@client.event
async def on_message(message):
    if message.author.bot: return
    content=message.content.strip(); uid=str(message.author.id)
    cid=str(message.channel.id)
    
    # 更新房間活動
    if cid in bot_data.get("rooms",{}):
        bot_data["rooms"][cid]["last_activity"]=datetime.now().isoformat(); save_data()
    
    # 快捷報班
    if content.startswith('/原推') or content.startswith('/s6') or content.startswith('/雙') or content.startswith('/三開'):
        if uid not in bot_data.get("members",{}):
            img=render_message_box("錯誤",["請先 /成員 註冊"],accent_color=Theme.RED)
            await message.reply(file=discord.File(img,"e.png"),silent=True); return
        if not bot_data.get("settings",{}).get("schedule_open"):
            img=render_message_box("錯誤",["報班未開放"],accent_color=Theme.RED)
            await message.reply(file=discord.File(img,"e.png"),silent=True); return
        parts=content.split()
        if len(parts)<2:
            img=render_message_box("格式",["如: /原推 08-12"],accent_color=Theme.ORANGE)
            await message.reply(file=discord.File(img,"fmt.png"),silent=True); return
        cmd,time_str=parts[0],parts[1]; note=" ".join(parts[2:]) if len(parts)>2 else ""
        role="s6" if cmd=='/s6' else "pusher"
        hours=parse_time_range(time_str)
        if not hours:
            img=render_message_box("錯誤",["格式: 08-12"],accent_color=Theme.RED)
            await message.reply(file=discord.File(img,"e.png"),silent=True); return
        today=get_today(); bot_data.setdefault("schedule",{}).setdefault(today,{})
        m=bot_data["members"][uid]
        multi_map={'/雙':'雙開','/三開':'三開'}
        app={"user_id":uid,"name":m["name"],"bonus":m["bonus"],"bonus_2":m.get("bonus_2",0),
             "bonus_3":m.get("bonus_3",0),"s6_bonus":m.get("s6_bonus",0),"power":m["power"],
             "s6_power":m.get("s6_power",0),"multi":multi_map.get(cmd,m["multi"]),
             "role":role,"note":note,"registered_at":datetime.now().isoformat()}
        registered=[]
        for h in hours:
            closed,_=is_signup_closed(h)
            if closed: continue
            if h not in bot_data["schedule"][today]: bot_data["schedule"][today][h]={"applicants":[]}
            if not any(a["user_id"]==uid for a in bot_data["schedule"][today][h].get("applicants",[])):
                bot_data["schedule"][today][h].setdefault("applicants",[]).append(app); registered.append(h)
        save_data()
        if registered:
            refresh_schedule(today)
            img=render_message_box("報班成功",[f"時段: {', '.join(registered)}",f"倍率: {m['bonus']:.2f}"],accent_color=Theme.GREEN)
            await message.reply(file=discord.File(img,"ok.png"),silent=True)
        return
    
    # 設定房號快捷
    if content.startswith('設定房號'):
        parts=content.split()
        if len(parts)>=3:
            room_id=parts[1]; car_type=parts[2] if parts[2] in CAR_TYPES else "蝦"
            orig=message.channel.name
            try: await message.channel.edit(name=f"{room_id}-{car_type}")
            except: pass
            bot_data["rooms"][cid]={"room_id":room_id,"car_type":car_type,"original_name":orig,
                "created_at":datetime.now().isoformat(),"last_activity":datetime.now().isoformat()}
            save_data()
            img=render_info_card("房間設定",[("房號",room_id),("車種",car_type)],accent_color=Theme.BLUE)
            await message.reply(file=discord.File(img,"room.png"),silent=True)
        return
    
    # 排名快捷: e50 / e1-10
    if content.lower().startswith('e') and len(content)>1 and content[1:2].isdigit():
        rank_part=content[1:].strip()
        try:
            async with ClientSession() as session:
                async with session.get(f"{HISEKAI_API}/event/live/top100",timeout=ClientTimeout(total=15)) as resp:
                    data=await resp.json()
            rankings=data.get('top_100_player_rankings',[]); event_name=data.get('name','-')
            if '-' in rank_part:
                parts=rank_part.split('-'); start=int(parts[0]); end=int(parts[1])
                if start>end: start,end=end,start
                img=create_ranking_list_image(rankings,max(1,start),min(100,end),event_name)
                if img: await message.reply(file=discord.File(img,f"rank_{start}_{end}.png"),silent=True)
            else:
                target_rank=int(rank_part)
                if target_rank<1 or target_rank>100: await message.reply("範圍: 1-100",silent=True); return
                target=prev_p=next_p=None
                for p in rankings:
                    if p.get('rank')==target_rank: target=p
                    if p.get('rank')==target_rank-1: prev_p=p
                    if p.get('rank')==target_rank+1: next_p=p
                if not target: await message.reply(f"找不到T{target_rank}",silent=True); return
                rk=str(target_rank); hd_list=[]
                cur=[r for r in ranking_history.get("records",[]) if r.get('event')==event_name]
                for rec in cur:
                    if rk in rec.get("borders",{}): hd_list.append({'time':rec['time'],'score':rec["borders"][rk]["score"]})
                img=create_ranking_detail_image(target,prev_p,next_p,event_name,hd_list)
                if img: await message.reply(file=discord.File(img,f"t{target_rank}.png"),silent=True)
        except ValueError: await message.reply("格式: e50 或 e1-10",silent=True)
        except Exception as e: await message.reply(f"查詢失敗: {e}",silent=True)
        return

# ========== 啟動 ==========
@client.event
async def on_ready():
    global table
    print(f"Bot: {client.user}")
    xlsx=os.path.join(os.path.dirname(os.path.abspath(__file__)),"score_data.xlsx")
    if os.path.exists(xlsx): table=ScoreTable(xlsx)
    await tree.sync(); print("Commands synced")
    client.loop.create_task(reminder_task())

if __name__=="__main__":
    token=os.getenv("DISCORD_BOT_TOKEN") or os.getenv("DISCORD_TOKEN")
    if not token: print("Error: DISCORD_BOT_TOKEN or DISCORD_TOKEN not set"); exit(1)
    client.run(token)
